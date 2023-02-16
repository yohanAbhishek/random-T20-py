import random
import openpyxl
import time


def toss(TeamID):  # Returns who won the toss

    while True:
        team1 = input(f"\n{TeamID[0:2]} Head or Tail: ").upper()
        if team1 == "HEAD":
            print(f"{TeamID[2:4]} is tails")
            break
        elif team1 == "TAIL":
            print(f"{TeamID[2:4]} is heads")
            break
        else:
            print("Invalid choice!")

    randToss = random.choice(["HEAD", "TAIL"])

    if team1 == randToss:
        tossWin = TeamID[0:2]
        tossLose = TeamID[2:4]
    else:
        tossWin = TeamID[2:4]
        tossLose = TeamID[0:2]
    print("\nChoosing randomly", end="")
    for i in range(3):
        print('.', end='')
        time.sleep(1)
    print(f"\n\nTeam {tossWin} won the toss!\n")

    while True:
        choice = input("Bat or bowl: ").lower()
        if choice == "bat":
            print(f"\nTeam {tossWin}: Batting\nTeam {tossLose}: Bowling")
            break
        elif choice == "bowl":
            print(f"Team {tossWin}: Bowling\nTeam {tossLose}: Batting")
            break
        else:
            print("Invalid choice!")
    return [tossWin, tossLose, choice]


batsman = fielder = ""


def randomAssign(groupBat, teamNoBat, groupField, teamNoField):
    global fielder, batsman
    fo = open(f"Storage/Group{groupBat}_profile/team_{teamNoBat}.txt", 'r')  # To select batsman
    team = fo.readlines()
    batsman = []
    for member in team:
        batsman.append(member[0:4])

    fo = open(f"Storage/Group{groupField}_profile/team_{teamNoField}.txt", 'r')  # To select bowler
    team = fo.readlines()
    fielder = []
    for member in team:
        fielder.append(member[0:4])

    random.shuffle(fielder)
    random.shuffle(batsman)


def saveDataBat(matchID, batsmanID, playerScore, ballsToScore, randomDismiss):
    wb = openpyxl.load_workbook("Storage/pointsTable.xlsx")
    sh1 = wb['Batsman']
    currentLine = sh1.max_row

    sh1.cell(row=currentLine + 1, column=1, value=matchID)  # matchID
    sh1.cell(row=currentLine + 1, column=2, value=batsmanID)  # batsmanID
    sh1.cell(row=currentLine + 1, column=3, value=playerScore)  # totalScore
    sh1.cell(row=currentLine + 1, column=4, value=ballsToScore)  # ballsToScore
    sh1.cell(row=currentLine + 1, column=5, value=randomDismiss[0])  # dismissal

    wb.save("Storage/pointsTable.xlsx")
    wb.close()


def saveDataBowl(matchID, bowlID, ballsBowled, wicketsTaken):
    wb = openpyxl.load_workbook("Storage/pointsTable.xlsx")
    sh1 = wb['Bowler']
    currentLine = sh1.max_row

    sh1.cell(row=currentLine + 1, column=1, value=matchID)  # matchID
    sh1.cell(row=currentLine + 1, column=2, value=bowlID)  # bowlID
    sh1.cell(row=currentLine + 1, column=3, value=ballsBowled)  # ballsBowled
    sh1.cell(row=currentLine + 1, column=4, value=wicketsTaken)  # wicketsTaken

    wb.save("Storage/pointsTable.xlsx")
    wb.close()


def executeGame(matchID, batTeam, bowlTeam):
    dismissMethod = ["Score", 'Caught', 'Bowled', 'LBW', 'Run-out', 'Stumped']
    randomAssign(batTeam[0], batTeam[1], bowlTeam[0], bowlTeam[1])  # assign a fielder to bowl against a batsman
    totalBalls = ballsToScore = matchWickets = playerScore = totalScore = wicketsTaken = bowler = ballsBowled = 0

    scores = [0, 1, 2, 3, 4, 6]

    bat = batsman[matchWickets]

    try:
        while totalBalls < 120 and matchWickets < 10:  # Full inning (20 overs per team)

            if totalBalls % 6 == 0 and totalBalls != 0:  # To change the bowler every over (6 balls)
                bowler = random.choice(fielder)
                saveDataBowl(matchID, bowler, ballsBowled, wicketsTaken)
                ballsBowled = 0
                wicketsTaken = 0
                print(f"\n>> {bowler} is bowling now\n")

            randomDismiss = random.choices(dismissMethod,
                                           weights=[random.randint(54, 55), 1, 1, 1, 1, 1])  # Return score or dismiss

            if randomDismiss == ["Score"]:
                ballsToScore += 1
                run = random.choices(scores, weights=[5, 5, 3, 1, 2, 1])  # Randomly select a score
                for i in run:
                    run = i
                totalScore += run  # Total match score
                playerScore += run  # Total player score

                print("Score:", run)

            else:  # If the batsman gets out
                ballsToScore += 1
                saveDataBat(matchID, bat, playerScore, ballsToScore, randomDismiss)
                print('\n' + "ID:", bat, "|Dismiss:", randomDismiss[0], '|', "Total player score:", playerScore)

                matchWickets += 1  # Total match wickets
                wicketsTaken += 1  # By each bowler

                if totalBalls < 120 and matchWickets < 10:
                    bat = batsman[matchWickets]  # Select a new batsman
                    print("\n" + bat, "is batting now..")
                    print('-' * 100)
                playerScore = 0
                ballsToScore = 0

            ballsBowled += 1
            totalBalls += 1

            if totalBalls >= 120:
                randomDismiss = ["Balls-over"]
                saveDataBat(matchID, bat, playerScore, ballsToScore, randomDismiss)
                saveDataBowl(matchID, bowler, ballsBowled, wicketsTaken)
                print('\n' + bat, randomDismiss[0], '|', "Total player score:", playerScore)

            elif matchWickets == 10:
                saveDataBowl(matchID, bowler, ballsBowled, wicketsTaken)

    except IndexError:
        print("\nAll players out!\n")
    print("\n>> Match over!\n")


def selectTeam():
    wb = openpyxl.load_workbook("Storage/pointsTable.xlsx")
    sh1 = wb["Batsman"]
    highest_row = sh1.max_row
    matches = []
    for i in range(2, highest_row + 1):  # Get the matchIDs of completed games
        match = sh1.cell(row=i, column=1).value
        if match not in matches:
            matches.append(match)

    matchCombination = {1: "A1 vs A2", 2: "A1 vs A3", 3: "A1 vs A4", 4: "A2 vs A3", 5: "A2 vs A4", 6: "A3 vs A4",
                        7: "B1 vs B2", 8: "B1 vs B3", 9: "B1 vs B4", 10: "B2 vs B3", 11: "B2 vs B4", 12: "B3 vs B4"}

    for match in range(1, 13):  # Print the teams
        if match == 1:
            print("\n-- Group A -- ")
        elif match == 7:
            print("\n-- Group B -- ")
        print(f"{match}. {matchCombination[match]}")

    while True:
        try:
            currentMatch = int(input("\nEnter the number from the above match combinations that you wish to start: "))
            machID = matchCombination[currentMatch][0:2] + matchCombination[currentMatch][6:8]
            if machID in matches:
                print('\nMatch already completed!')
            else:
                break
        except ValueError:
            print("Please enter a number!")
        except KeyError:
            print("Invalid ID!")

    print(f'\nStarting {matchCombination[currentMatch][0:2]} vs {matchCombination[currentMatch][6:8]}', end="")
    for i in range(3):
        print('.', end='')
        time.sleep(1)
    print()
    return machID
