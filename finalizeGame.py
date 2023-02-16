import openpyxl


def summary(matchID):
    team1 = matchID[0:2]
    team2 = matchID[2:4]

    team1TotalRuns = calcTotal(team1, matchID, 2, 1, 3, "Batsman")    # |
    team1TotalBalls = calcTotal(team1, matchID, 2, 1, 3, "Bowler")    # Calculate runs, balls, wickets for team1
    team1TotalWickets = calcTotal(team1, matchID, 2, 1, 4, "Bowler")  # |

    team2TotalRuns = calcTotal(team2, matchID, 2, 1, 3, "Batsman")    # |
    team2TotalBalls = calcTotal(team2, matchID, 2, 1, 3, "Bowler")    # Calculate runs, balls, wickets for team2
    team2TotalWickets = calcTotal(team2, matchID, 2, 1, 4, "Bowler")  # |

    status1 = selectWinner(team1TotalRuns, team2TotalRuns)[0]  # Find winner and looser
    status2 = selectWinner(team1TotalRuns, team2TotalRuns)[1]

    print("\n", " " * 35, "Match summary")  # Print the table of summary
    print("_" * 86)
    print("| {:15}| {:15}| {:15}| {:15}| {:15}|".format('Team', 'Total Score', 'Total balls', 'Wickets', 'Won/Lost'))
    print("_" * 86)

    print("| {:15}| {:15}| {:15}| {:15}| {:15}|".format(team1, team1TotalRuns, team1TotalBalls, team1TotalWickets,
                                                        status1))
    print("| {:15}| {:15}| {:15}| {:15}| {:15}|".format(team2, team2TotalRuns, team2TotalBalls, team2TotalWickets,
                                                        status2))
    print("-" * 86)


def highestScoredBatsman():
    wb = openpyxl.load_workbook("Storage/pointsTable.xlsx")
    sh1 = wb["Batsman"]
    playerScores = []
    unsorted = []
    highestFive = []  # row number of each batsman
    highestFiveID = []
    highestFiveScores = []
    p1Name = p2Name = p3Name = p4Name = p5Name = ''

    highestRow = sh1.max_row
    if highestRow > 1:
        for i in range(2, highestRow + 1):  # get scores and store in playerScores
            score = sh1.cell(row=i, column=3).value
            playerScores.append(score)
            unsorted.append(score)

        playerScores.sort()
        sortedScores = playerScores[::-1]  # Sort in descending

        for i in range(5):  # get five highest numbers(row numbers)
            highestFive.append(unsorted.index(sortedScores[i]) + 2)

        for i in highestFive:  # Get ID and score with reference to above three numbers(rows)
            batsmanScore = sh1.cell(row=i, column=3).value
            batsmanID = sh1.cell(row=i, column=2).value
            highestFiveID.append(batsmanID)
            highestFiveScores.append(batsmanScore)

        highest1 = [highestFiveID[0], highestFiveScores[0]]
        highest2 = [highestFiveID[1], highestFiveScores[1]]
        highest3 = [highestFiveID[2], highestFiveScores[2]]
        highest4 = [highestFiveID[3], highestFiveScores[3]]
        highest5 = [highestFiveID[4], highestFiveScores[4]]

        for item in ['A', 'B']:
            for num in [1, 2, 3, 4]:
                fo = open(f"Storage/Group{item}_profile/team_{num}.txt", "r")  # Read one team

                a = fo.readlines()
                for i in a:
                    if highest1[0] in i:
                        p1Name = i.split(',')[1]
                    if highest2[0] in i:
                        p2Name = i.split(',')[1]
                    if highest3[0] in i:
                        p3Name = i.split(',')[1]
                    if highest4[0] in i:
                        p4Name = i.split(',')[1]
                    if highest5[0] in i:
                        p5Name = i.split(',')[1]

                fo.close()

        # Print top 5 batsman in table format
        print("\n", " " * 25, "Top 5 batsman")
        print("_" * 67)
        print("| {:15}| {:30}| {:15}|".format('Player ID', 'Player name', 'Score'))
        print("_" * 67)

        print("| {:15}| {:30}| {:15}|".format(highest1[0], p1Name, highest1[1]))
        print("| {:15}| {:30}| {:15}|".format(highest2[0], p2Name, highest2[1]))
        print("| {:15}| {:30}| {:15}|".format(highest3[0], p3Name, highest3[1]))
        print("| {:15}| {:30}| {:15}|".format(highest4[0], p4Name, highest4[1]))
        print("| {:15}| {:30}| {:15}|".format(highest5[0], p5Name, highest5[1]))
        print("-" * 67)
    else:
        print('>> No games played yet')


def highestWicketBowlers():
    wb = openpyxl.load_workbook("Storage/pointsTable.xlsx")
    sh1 = wb["Bowler"]
    highestRow = sh1.max_row
    teamID = []
    wicketBowlerDict = {}
    total = 0
    p1Name = p2Name = p3Name = p4Name = p5Name = ''

    if highestRow > 1:
        for i in range(2, highestRow + 1):  # get teamID of all bowlers
            ID = sh1.cell(row=i, column=2).value
            if ID not in teamID:
                teamID.append(ID)

        for i in teamID:  # Locate bowler in excel and calculate total wickets for each bowler
            for a in range(2, highestRow):
                ID = sh1.cell(row=a, column=2).value
                if i == ID:
                    wicket = sh1.cell(row=a, column=4).value
                    total += wicket
            wicketBowlerDict[i] = total
            total = 0

        highestFive = sorted(wicketBowlerDict, key=wicketBowlerDict.get, reverse=True)[:5]

        # Printing the top 5 bowlers based on wickets taken
        highest1 = [highestFive[0], wicketBowlerDict[highestFive[0]]]
        highest2 = [highestFive[1], wicketBowlerDict[highestFive[1]]]
        highest3 = [highestFive[2], wicketBowlerDict[highestFive[2]]]
        highest4 = [highestFive[3], wicketBowlerDict[highestFive[3]]]
        highest5 = [highestFive[4], wicketBowlerDict[highestFive[4]]]

        for item in ['A', 'B']:
            for num in [1, 2, 3, 4]:
                fo = open(f"Storage/Group{item}_profile/team_{num}.txt", "r")  # Read one team

                a = fo.readlines()
                for i in a:
                    if highest1[0] in i:
                        p1Name = i.split(',')[1]
                    if highest2[0] in i:
                        p2Name = i.split(',')[1]
                    if highest3[0] in i:
                        p3Name = i.split(',')[1]
                    if highest4[0] in i:
                        p4Name = i.split(',')[1]
                    if highest5[0] in i:
                        p5Name = i.split(',')[1]

                fo.close()

        # Print top 5 batsman in table format
        print("\n", " " * 25, "Top 5 bowlers")
        print("_" * 67)
        print("| {:15}| {:30}| {:15}|".format('Player ID', 'Player name', 'Wickets'))
        print("_" * 67)

        print("| {:15}| {:30}| {:15}|".format(highest1[0], p1Name, highest1[1]))
        print("| {:15}| {:30}| {:15}|".format(highest2[0], p2Name, highest2[1]))
        print("| {:15}| {:30}| {:15}|".format(highest3[0], p3Name, highest3[1]))
        print("| {:15}| {:30}| {:15}|".format(highest4[0], p4Name, highest4[1]))
        print("| {:15}| {:30}| {:15}|".format(highest5[0], p5Name, highest5[1]))
        print("-" * 67)
    else:
        print('>> No games played yet')


def tournamentStanding():
    wb = openpyxl.load_workbook("Storage/pointsTable.xlsx")
    sh1 = wb["Batsman"]
    highestRow = sh1.max_row

    winnings = {'A1': [0, 0], 'A2': [0, 0], 'A3': [0, 0], 'A4': [0, 0], 'B1': [0, 0], 'B2': [0, 0], 'B3': [0, 0],
                'B4': [0, 0]}
    matches = []
    for i in range(2, highestRow + 1):  # Get the matchIDs of completed games
        match = sh1.cell(row=i, column=1).value
        if match not in matches:
            matches.append(match)

    for matchID in matches:
        team1 = matchID[0:2]
        team2 = matchID[2:4]
        team1TotalRuns = calcTotal(team1, matchID, 2, 1, 3, "Batsman")
        team2TotalRuns = calcTotal(team2, matchID, 2, 1, 3, "Batsman")

        status1 = selectWinner(team1TotalRuns, team2TotalRuns)[0]  # Find winner and looser
        status2 = selectWinner(team1TotalRuns, team2TotalRuns)[1]

        if status1 == "Won":  # Save win/loss by the two teams
            winnings[team1] = [winnings[team1][0] + 1, winnings[team1][1]]
            winnings[team2] = [winnings[team2][0], winnings[team2][1] + 1]

        elif status2 == "Won":
            winnings[team2] = [winnings[team2][0] + 1, winnings[team2][1]]
            winnings[team1] = [winnings[team1][0], winnings[team1][1] + 1]
        # If draw, it is not considered

    print("_" * 37)
    print("| {:10}| {:10}| {:10}|".format(' TeamID', '   Won', '  Lost'))
    print("-" * 37)

    for i in winnings:
        print("| {:10}| {:10}| {:10}|".format(f"    {i}", f"    {winnings[i][0]}", f"    {winnings[i][1]}"))
    print("-" * 37)


def selectIdForSummary():

    matchCombination = {1: "A1 vs A2", 2: "A1 vs A3", 3: "A1 vs A4", 4: "A2 vs A3", 5: "A2 vs A4", 6: "A3 vs A4",
                        7: "B1 vs B2", 8: "B1 vs B3", 9: "B1 vs B4", 10: "B2 vs B3", 11: "B2 vs B4", 12: "B3 vs B4"}
    for match in range(1, 13):
        if match == 1:
            print("\n-- Group A -- ")
        elif match == 7:
            print("\n-- Group B -- ")
        print(f"{match}. {matchCombination[match]}")
    while True:
        try:
            currentMatch = int(input("\nEnter option: "))
            if currentMatch not in matchCombination:
                print("\nOption Invalid!")
            else:
                matchID = matchCombination[currentMatch][0:2] + matchCombination[currentMatch][6:8]
                return matchID
        except ValueError:
            print('Please enter a number!')


def calcTotal(teamNo, matchID, column1, column2, column3, excelFile):
    wb = openpyxl.load_workbook("Storage/pointsTable.xlsx")
    sh1 = wb[excelFile]
    highestRow = sh1.max_row
    storeRuns = 0
    for i in range(2, highestRow + 1):  # Get total score for team1
        playerID = sh1.cell(row=i, column=column1).value
        match = sh1.cell(row=i, column=column2).value
        if teamNo in playerID and match == matchID:
            score = sh1.cell(row=i, column=column3).value
            storeRuns += int(score)
    return storeRuns


def selectWinner(team1TotalRuns, team2TotalRuns):

    if team1TotalRuns > team2TotalRuns:  # Compare runs of both teams and select winner/ draw
        status1 = "Won"
        status2 = "Lost"
    elif team1TotalRuns < team2TotalRuns:
        status1 = "Won"
        status2 = "Lost"
    elif team1TotalRuns == team2TotalRuns and team2TotalRuns != 0:
        status1 = "Draw"
        status2 = "Draw"
    else:
        status1 = "Not played"
        status2 = "Not played"

    return [status1, status2]


def checkEnd():
    wb = openpyxl.load_workbook("Storage/pointsTable.xlsx")
    sh1 = wb["Batsman"]
    highestRow = sh1.max_row
    teams = []
    matches = ['A1A2', 'A1A3', 'A1A4', 'A2A3', 'A2A4', 'A3A4', 'B1B2', 'B1B3', 'B1B4', 'B2B3', 'B2B4',  'B3B4']
    teamsASCII = 0
    matchesASCII = 0

    for i in range(2, highestRow + 1):
        match = sh1.cell(row=i, column=1).value
        if match not in teams:
            teams.append(match)

    for i in matches:
        total = sum(ord(each) for each in i)
        matchesASCII += total

    for i in teams:
        total = sum(ord(each) for each in i)
        teamsASCII += total

    if teamsASCII == matchesASCII:  # If all matches have been played
        return True

    else:
        return False
